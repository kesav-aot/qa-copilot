"""Parsers for the formats QA teams actually keep test cases in.

Each returns a list of :class:`ManualTestCase`. Parsers are forgiving by design —
real test cases are inconsistently written, and it is the analyser's job to say
what is missing, not the parser's job to reject the file.
"""

from __future__ import annotations

import csv
import io
import json
import re
from pathlib import Path

from qa_copilot.ingest.models import ManualStep, ManualTestCase

# ``1. do the thing``  /  ``1) do the thing``  /  ``- do the thing``
_NUMBERED = re.compile(r"^\s*(?:(\d+)[.)]|[-*•])\s+(.*)$")
# ``do the thing -> expected result``  and the ``|`` / ``=>`` variants
_INLINE_EXPECTED = re.compile(r"\s*(?:->|=>|\|)\s*")
_ID_IN_TITLE = re.compile(r"\b([A-Z][A-Z0-9]*-\d+|TC[-_]?\d+)\b")


def _slug_id(text: str, fallback: str) -> str:
    match = _ID_IN_TITLE.search(text)
    if match:
        return match.group(1).upper()
    cleaned = re.sub(r"[^A-Za-z0-9]+", "-", text).strip("-").lower()
    return (cleaned or fallback)[:48]


def _split_steps(blob: str) -> list[ManualStep]:
    """Turn a free-text step block into numbered steps, keeping any inline
    expectation attached to its step."""
    steps: list[ManualStep] = []
    for raw in (blob or "").splitlines():
        line = raw.strip()
        if not line:
            continue
        match = _NUMBERED.match(line)
        body = match.group(2).strip() if match else line
        if not body:
            continue
        action, expected = body, None
        if _INLINE_EXPECTED.search(body):
            action, expected = (p.strip() for p in _INLINE_EXPECTED.split(body, maxsplit=1))
        steps.append(ManualStep(number=len(steps) + 1, action=action, expected=expected or None))
    return steps


def _split_lines(blob: str) -> list[str]:
    out = []
    for raw in (blob or "").replace(";", "\n").splitlines():
        line = raw.strip()
        match = _NUMBERED.match(line)
        if match:
            line = match.group(2).strip()
        if line:
            out.append(line)
    return out


# --- markdown --------------------------------------------------------------

_SECTIONS = {
    "precondition": "preconditions",
    "preconditions": "preconditions",
    "prerequisite": "preconditions",
    "prerequisites": "preconditions",
    "setup": "preconditions",
    "given": "preconditions",
    "step": "steps",
    "steps": "steps",
    "procedure": "steps",
    "actions": "steps",
    "test steps": "steps",
    "expected": "expected",
    "expected result": "expected",
    "expected results": "expected",
    "expected outcome": "expected",
    "acceptance criteria": "expected",
    "description": "description",
    "summary": "description",
    "objective": "description",
    "tags": "tags",
    "labels": "tags",
    "priority": "priority",
}


def parse_markdown(text: str, source: str) -> list[ManualTestCase]:
    """Test cases separated by ``#``/``##`` headings, with labelled sections.

    Also accepts ``**Steps:**`` style bold labels and ``Steps:`` plain labels,
    because exported test cases use all three.
    """
    cases: list[ManualTestCase] = []
    blocks = re.split(r"^#{1,3}\s+", text, flags=re.MULTILINE)
    for index, block in enumerate(blocks):
        if not block.strip():
            continue
        lines = block.splitlines()
        heading = lines[0].strip()
        if not heading:
            continue

        buckets: dict[str, list[str]] = {}
        current = "description"
        for line in lines[1:]:
            stripped = line.strip().strip("*_")
            label = re.match(r"^([A-Za-z][A-Za-z ]{2,24})\s*:\s*(.*)$", stripped)
            if label and label.group(1).strip().lower() in _SECTIONS:
                current = _SECTIONS[label.group(1).strip().lower()]
                remainder = label.group(2).strip()
                if remainder:
                    buckets.setdefault(current, []).append(remainder)
                continue
            buckets.setdefault(current, []).append(line)

        title = re.sub(r"^\s*(?:[A-Z][A-Z0-9]*-\d+|TC[-_]?\d+)\s*[:\-–]\s*", "", heading).strip()
        cases.append(
            ManualTestCase(
                id=_slug_id(heading, f"case-{index}"),
                title=title or heading,
                source=source,
                format="markdown",
                description=" ".join(_split_lines("\n".join(buckets.get("description", [])))) or None,
                preconditions=_split_lines("\n".join(buckets.get("preconditions", []))),
                steps=_split_steps("\n".join(buckets.get("steps", []))),
                expected_results=_split_lines("\n".join(buckets.get("expected", []))),
                tags=[t for t in re.split(r"[,\s]+", " ".join(buckets.get("tags", []))) if t],
                priority=(" ".join(buckets.get("priority", [])).strip() or None),
            )
        )
    return cases


# --- csv / excel -----------------------------------------------------------

_COLUMN_ALIASES = {
    "id": {"id", "key", "test id", "testcase id", "test case id", "case id", "tc", "tc id"},
    "title": {"title", "summary", "name", "test case", "test name", "scenario"},
    "description": {"description", "objective", "notes"},
    "preconditions": {"precondition", "preconditions", "prerequisite", "prerequisites", "setup"},
    "steps": {"step", "steps", "test steps", "procedure", "actions", "action"},
    "expected": {"expected", "expected result", "expected results", "expected outcome", "result"},
    "tags": {"tags", "labels", "component", "components"},
    "priority": {"priority", "severity"},
}


def _map_columns(headers: list[str]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for header in headers:
        key = (header or "").strip().lower()
        for field, aliases in _COLUMN_ALIASES.items():
            if key in aliases and field not in mapping:
                mapping[field] = header
                break
    return mapping


def _rows_to_cases(rows: list[dict], source: str, fmt: str) -> list[ManualTestCase]:
    if not rows:
        return []
    mapping = _map_columns(list(rows[0].keys()))
    if "title" not in mapping and "steps" not in mapping:
        raise ValueError(
            f"{source}: could not find a title or steps column. "
            f"Columns seen: {', '.join(str(c) for c in rows[0])}"
        )

    cases: list[ManualTestCase] = []
    for index, row in enumerate(rows):
        def cell(field: str) -> str:
            column = mapping.get(field)
            value = row.get(column) if column else None
            return "" if value is None else str(value).strip()

        title = cell("title") or f"Untitled case {index + 1}"
        if not any(cell(f) for f in ("title", "steps", "expected")):
            continue
        cases.append(
            ManualTestCase(
                id=cell("id") or _slug_id(title, f"row-{index + 1}"),
                title=title,
                source=source,
                format=fmt,  # type: ignore[arg-type]
                description=cell("description") or None,
                preconditions=_split_lines(cell("preconditions")),
                steps=_split_steps(cell("steps")),
                expected_results=_split_lines(cell("expected")),
                tags=[t for t in re.split(r"[,;\s]+", cell("tags")) if t],
                priority=cell("priority") or None,
            )
        )
    return cases


def parse_csv(text: str, source: str) -> list[ManualTestCase]:
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    rows = list(csv.DictReader(io.StringIO(text), dialect=dialect))
    return _rows_to_cases(rows, source, "csv")


def parse_excel(path: Path) -> list[ManualTestCase]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - optional dependency
        raise RuntimeError(
            "reading .xlsx needs openpyxl: pip install 'qa-copilot[excel]'"
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    rows: list[dict] = []
    for sheet in workbook.worksheets:
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            continue
        headers = [str(h) if h is not None else "" for h in values[0]]
        for row in values[1:]:
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            rows.append(dict(zip(headers, row, strict=False)))
    workbook.close()
    return _rows_to_cases(rows, str(path), "excel")


# --- gherkin ---------------------------------------------------------------

def parse_gherkin(text: str, source: str) -> list[ManualTestCase]:
    """Given → preconditions, When/And → steps, Then → expected results."""
    cases: list[ManualTestCase] = []
    feature = ""
    current: dict | None = None
    mode = ""

    def flush() -> None:
        if current and (current["steps"] or current["expected"]):
            title = current["title"]
            cases.append(
                ManualTestCase(
                    id=_slug_id(title, f"scenario-{len(cases) + 1}"),
                    title=title,
                    source=source,
                    format="gherkin",
                    description=feature or None,
                    preconditions=current["given"],
                    steps=[
                        ManualStep(number=i + 1, action=a)
                        for i, a in enumerate(current["steps"])
                    ],
                    expected_results=current["expected"],
                    tags=current["tags"],
                )
            )

    pending_tags: list[str] = []
    for raw in text.splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        if line.startswith("@"):
            pending_tags = [t.lstrip("@") for t in line.split()]
            continue
        if line.lower().startswith("feature:"):
            feature = line.split(":", 1)[1].strip()
            continue
        if re.match(r"(?i)^(scenario outline|scenario|example):", line):
            flush()
            current = {
                "title": line.split(":", 1)[1].strip(),
                "given": [],
                "steps": [],
                "expected": [],
                "tags": pending_tags,
            }
            pending_tags = []
            mode = ""
            continue
        if current is None:
            continue

        keyword = re.match(r"(?i)^(given|when|then|and|but)\b\s*(.*)$", line)
        if keyword:
            word, body = keyword.group(1).lower(), keyword.group(2).strip()
            if word in {"given", "when", "then"}:
                mode = word
            bucket = {"given": "given", "when": "steps", "then": "expected"}.get(mode)
            if bucket == "given":
                current["given"].append(body)
            elif bucket == "expected":
                current["expected"].append(body)
            else:
                current["steps"].append(body)
    flush()
    return cases


# --- jira export -----------------------------------------------------------

def _adf_to_text(node) -> str:
    """Flatten Atlassian Document Format to plain text."""
    if isinstance(node, str):
        return node
    if isinstance(node, list):
        return "\n".join(_adf_to_text(n) for n in node)
    if isinstance(node, dict):
        if node.get("type") == "text":
            return node.get("text", "")
        inner = _adf_to_text(node.get("content", []))
        if node.get("type") in {"paragraph", "listItem", "heading"}:
            return inner + "\n"
        return inner
    return ""


def parse_jira(text: str, source: str) -> list[ManualTestCase]:
    """A Jira REST search export: ``{"issues": [...]}``, or a bare list."""
    payload = json.loads(text)
    issues = payload.get("issues", payload) if isinstance(payload, dict) else payload
    if not isinstance(issues, list):
        raise ValueError(f"{source}: expected a Jira export with an 'issues' array")

    cases: list[ManualTestCase] = []
    for index, issue in enumerate(issues):
        fields = issue.get("fields", issue) if isinstance(issue, dict) else {}
        key = issue.get("key") if isinstance(issue, dict) else None
        summary = fields.get("summary") or f"Untitled issue {index + 1}"
        description = fields.get("description")
        body = _adf_to_text(description) if not isinstance(description, str) else description

        # Reuse the markdown section splitter on the description body.
        sub = parse_markdown(f"# {summary}\n{body or ''}", source)
        case = sub[0] if sub else ManualTestCase(
            id=key or _slug_id(summary, f"issue-{index}"),
            title=summary,
            source=source,
            format="jira",
        )
        labels = fields.get("labels") or []
        priority = fields.get("priority")
        cases.append(
            case.model_copy(
                update={
                    "id": key or case.id,
                    "title": summary,
                    "format": "jira",
                    "source": source,
                    "tags": [str(t) for t in labels],
                    "priority": (priority or {}).get("name") if isinstance(priority, dict) else priority,
                }
            )
        )
    return cases


# --- plain text ------------------------------------------------------------

def parse_text(text: str, source: str) -> list[ManualTestCase]:
    """Last resort: treat the whole file as one case, using the section labels
    the markdown parser understands."""
    first = next((line.strip() for line in text.splitlines() if line.strip()), "Untitled")
    cases = parse_markdown(f"# {first}\n" + text.split(first, 1)[-1], source)
    return [c.model_copy(update={"format": "text"}) for c in cases]
